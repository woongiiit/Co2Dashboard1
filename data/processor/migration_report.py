#!/usr/bin/env python3
from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
OLD = ROOT / "data/excel/region/대시보드_임의데이터_확장_탄소발자국지수(260605).xlsx"
NEW = ROOT / "data/excel/region/★최종★탄소발자국_수식_산정(시안용).xlsx"
OUT = ROOT / "data/processor/migration_report.json"

INDUSTRIES = [
    "기타관광쇼핑",
    "대형쇼핑몰",
    "레저용품쇼핑",
    "면세점",
    "기타숙박",
    "캠핑장/펜션",
    "콘도",
    "호텔",
    "일반외식업",
    "제과음료업",
    "골프장",
    "관광유원시설",
    "기타레저",
    "문화서비스",
    "스키장",
    "카지노",
    "여행업",
    "렌터카",
    "수상운송",
    "육상운송",
    "항공운송",
    "뷰티",
    "의료관광",
]

C_BLOCK = "C: 에너지보완(mg)\n=B×보정계수"
IDX_KG = 153
IDX_INDEX = 154


def parse_num(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def build_composites(header_rows: list[tuple]) -> list[str]:
    r1 = [str(x).strip() if x is not None else "" for x in header_rows[0]]
    r2 = [str(x).strip() if x is not None else "" for x in header_rows[1]]
    composites: list[str] = []
    current = ""
    for group, sub in zip(r1, r2):
        if group:
            current = group
        if sub:
            composites.append(f"{current}||{sub}")
        elif current:
            composites.append(current)
        else:
            composites.append("")
    return composites


def load_old() -> dict[tuple, dict]:
    wb = openpyxl.load_workbook(OLD, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()
    headers = [str(x).strip() if x is not None else "" for x in rows[0]]
    col = {name: idx for idx, name in enumerate(headers) if name}
    out: dict[tuple, dict] = {}
    for row in rows[1:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        key = (
            str(row[col["sido_nm"]]).strip(),
            str(row[col["sgg_nm"]]).strip(),
            int(parse_num(row[col["year"]])),
            int(parse_num(row[col["month"]])),
        )
        out[key] = {
            "carbon_t": parse_num(row[col["탄소배출량"]]),
            "index": parse_num(row[col["탄소발자국 지수"]]),
            "industries_t": {name: parse_num(row[col[name]]) for name in INDUSTRIES if name in col},
        }
    return out


def load_new() -> tuple[list[str], dict[tuple, dict]]:
    wb = openpyxl.load_workbook(NEW, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()
    composites = build_composites(rows[:2])
    col = {name: idx for idx, name in enumerate(composites) if name}
    c_cols = {
        name: col[f"{C_BLOCK}||{name}"]
        for name in INDUSTRIES
        if f"{C_BLOCK}||{name}" in col
    }
    out: dict[tuple, dict] = {}
    for row in rows[2:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        key = (
            str(row[col["sido_nm"]]).strip(),
            str(row[col["sgg_nm"]]).strip(),
            int(parse_num(row[col["year"]])),
            int(parse_num(row[col["month"]])),
        )
        carbon_kg = parse_num(row[IDX_KG])
        out[key] = {
            "carbon_kg": carbon_kg,
            "carbon_t": carbon_kg / 1000.0,
            "index": parse_num(row[IDX_INDEX]),
            "industries_t": {
                name: parse_num(row[idx]) / 1_000_000_000.0 for name, idx in c_cols.items()
            },
        }
    return composites, out


def pct_diff(old: float, new: float) -> float | None:
    if old == 0:
        return None if new == 0 else float("inf")
    return (new - old) / old


def main() -> None:
    old = load_old()
    composites, new = load_new()

    common = set(old) & set(new)
    c_cols_found = [n for n in INDUSTRIES if f"{C_BLOCK}||{n}" in {c for c in composites}]

    carbon_diff_rows = 0
    index_diff_rows = 0
    carbon_samples: list[dict] = []
    index_samples: list[dict] = []
    industry_diff_by_name: dict[str, int] = defaultdict(int)

    for key in common:
        o = old[key]
        n = new[key]
        rel = pct_diff(o["carbon_t"], n["carbon_t"])
        if abs(o["carbon_t"] - n["carbon_t"]) > max(1.0, abs(o["carbon_t"]) * 0.005):
            carbon_diff_rows += 1
            if len(carbon_samples) < 8:
                carbon_samples.append(
                    {
                        "key": key,
                        "old_t": round(o["carbon_t"], 3),
                        "new_t": round(n["carbon_t"], 3),
                        "delta_pct": round(rel * 100, 2) if rel not in (None, float("inf")) else rel,
                    }
                )
        if abs(o["index"] - n["index"]) > 0.05:
            index_diff_rows += 1
            if len(index_samples) < 5:
                index_samples.append(
                    {
                        "key": key,
                        "old": round(o["index"], 4),
                        "new": round(n["index"], 4),
                    }
                )
        for name in c_cols_found:
            ov = o["industries_t"].get(name, 0.0)
            nv = n["industries_t"].get(name, 0.0)
            if abs(ov - nv) > max(0.1, abs(ov) * 0.01):
                industry_diff_by_name[name] += 1

    old_carbon_sum = sum(v["carbon_t"] for v in old.values())
    new_carbon_sum = sum(v["carbon_t"] for v in new.values())
    old_ind_sum = {n: sum(v["industries_t"].get(n, 0) for v in old.values()) for n in INDUSTRIES}
    new_ind_sum = {
        n: sum(v["industries_t"].get(n, 0) for v in new.values()) for n in c_cols_found
    }

    # converter compatibility
    converter_issues = []
    if "탄소배출량" not in composites and not any("탄소배출량" in c for c in composites):
        converter_issues.append("탄소배출량 컬럼 없음 → H 총(kg) 컬럼 매핑 필요")
    if not any("탄소발자국 지수" in c for c in composites):
        converter_issues.append("탄소발자국 지수 컬럼 없음 → H지수 컬럼 매핑 필요")
    converter_issues.append("2행 헤더(그룹+업종) 구조 → convert_region_excel.py 1행 헤더 가정과 불일치")
    missing_ind = [n for n in INDUSTRIES if n not in c_cols_found]
    if missing_ind:
        converter_issues.append(f"업종 컬럼 누락: {', '.join(missing_ind)}")

    # test current converter on new file
    try:
        import convert_region_excel as conv

        wb = openpyxl.load_workbook(NEW, read_only=True, data_only=True)
        conv.convert_workbook(wb.active)
        wb.close()
        converter_result = "success_unexpected"
    except Exception as exc:
        converter_result = f"{type(exc).__name__}: {exc}"

    # which file converter picks by default (sorted glob)
    picked = sorted(OLD.parent.glob("*.xlsx"))[0].name

    report = {
        "files": {
            "old": OLD.name,
            "new": NEW.name,
            "converter_default_pick": picked,
        },
        "structure": {
            "old": {
                "sheet": "Sheet1",
                "header_rows": 1,
                "columns": 31,
                "industry_columns": 23,
                "carbon_field": "탄소배출량 (tCO₂eq 추정)",
                "index_field": "탄소발자국 지수",
            },
            "new": {
                "sheet": "탄소발자국_데이터셋",
                "header_rows": 2,
                "columns": 155,
                "pipeline": [
                    "카드 사용량",
                    "탄소배출계수",
                    "중분류 업종별 가중치",
                    "에너지 조정승수",
                    "철도 배출",
                    "A 업종별배출량(mg)",
                    "B 가중치보완(mg)",
                    "C 에너지보완(mg)",
                    "D 6대분류(mg)",
                    "E/F/G/H 합산",
                ],
                "carbon_field": "H 총(kg) → tCO₂eq는 /1000",
                "index_field": "H지수 (연평균=100)",
                "industry_source_recommended": "C 에너지보완(mg) → tCO₂eq는 /1e9",
                "industry_columns_in_C_block": len(c_cols_found),
            },
        },
        "coverage": {
            "old_rows": len(old),
            "new_rows": len(new),
            "common_keys": len(common),
            "only_in_old": len(set(old) - set(new)),
            "only_in_new": len(set(new) - set(old)),
            "ym_range_old": [min(k[2:4] for k in old), max(k[2:4] for k in old)],
            "ym_range_new": [min(k[2:4] for k in new), max(k[2:4] for k in new)],
            "regions_old": len({(k[0], k[1]) for k in old}),
            "regions_new": len({(k[0], k[1]) for k in new}),
        },
        "missing_in_new": {
            "industry_columns": [n for n in INDUSTRIES if n not in c_cols_found],
        },
        "totals_comparison": {
            "carbon_sum_old_t": round(old_carbon_sum, 2),
            "carbon_sum_new_t": round(new_carbon_sum, 2),
            "carbon_sum_delta_pct": round((new_carbon_sum - old_carbon_sum) / old_carbon_sum * 100, 2)
            if old_carbon_sum
            else None,
            "industry_totals_delta_pct": {
                name: round((new_ind_sum.get(name, 0) - old_ind_sum.get(name, 0)) / old_ind_sum[name] * 100, 2)
                if old_ind_sum.get(name)
                else None
                for name in c_cols_found
                if abs(old_ind_sum.get(name, 0) - new_ind_sum.get(name, 0)) > 1
            },
        },
        "row_level_diffs_on_common_keys": {
            "carbon_rows_differ_gt_0_5pct": carbon_diff_rows,
            "index_rows_differ_gt_0_05": index_diff_rows,
            "carbon_samples": carbon_samples,
            "index_samples": index_samples,
            "industry_rows_differ_by_name": dict(industry_diff_by_name),
        },
        "converter": {
            "current_script_on_new_file": converter_result,
            "issues": converter_issues,
        },
        "service_impact": [],
    }

    # derive impact bullets
    impacts = report["service_impact"]
    impacts.append("convert_region_excel.py 수정 없이는 신규 파일 변환 불가")
    if missing_ind:
        impacts.append("신규 파일에 '여행업' 중분류 컬럼 없음 → 업종/지역 KPI 합산 불일치 가능")
    if carbon_diff_rows > len(common) * 0.1:
        impacts.append("동일 키 기준 탄소 총량 값이 대량 불일치 → KPI·지도·순위 수치 변경")
    if index_diff_rows > len(common) * 0.1:
        impacts.append("탄소발자국 지수 값 변경 → 지표/비교 로직 영향")
    if picked == NEW.name:
        impacts.append("폴더 내 xlsx 정렬상 신규 파일이 converter 기본 입력으로 선택될 위험")
    elif picked == OLD.name:
        impacts.append("현재 converter 기본 입력은 구 파일(정렬상 첫 번째)")

    OUT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({k: report[k] for k in ("files", "coverage", "missing_in_new", "totals_comparison", "converter")}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
