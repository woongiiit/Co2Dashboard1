#!/usr/bin/env python3
"""Compare old vs new region Excel files for migration assessment."""

from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl required", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[2]
REGION_DIR = ROOT / "data" / "excel" / "region"

OLD = REGION_DIR / "대시보드_임의데이터_확장_탄소발자국지수(260605).xlsx"
NEW = REGION_DIR / "★최종★탄소발자국_수식_산정(시안용).xlsx"

INDUSTRY_COLUMNS = [
    "기타관광쇼핑",
    "대형쇼핑몰",
    "레저용품쇼핑",
    "면세점",
    "기타숙박",
    "캠핑장/펜션",
    "콘도",
    "호텔",
    "일반외식업",
    "제과음료업",
    "골프장",
    "관광유원시설",
    "기타레isure",
]

# fix typo - use full list from converter
INDUSTRY_COLUMNS = [
    "기타관광쇼핑",
    "대형쇼핑몰",
    "레저용품쇼핑",
    "면세점",
    "기타숙박",
    "캠핑장/펜션",
    "콘도",
    "호텔",
    "일반외식업",
    "제과음료업",
    "골프장",
    "관광유원시설",
    "기타레저",
    "문화서비스",
    "스키장",
    "카지노",
    "여행업",
    "렌터카",
    "수상운송",
    "육상운송",
    "항공운송",
    "뷰티",
    "의료관광",
]

REQUIRED = ["sido_nm", "sgg_nm", "year", "month", "탄소발자국 지수"]
CARBON_HEADERS = ("탄소배출량 ", "탄소배출량")


def parse_number(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def load_rows(path: Path) -> tuple[list[str], list[tuple]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    wb.close()
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    data = [r for r in rows[1:] if r and any(c is not None and str(c).strip() != "" for c in r)]
    return headers, data


def carbon_idx(headers: list[str]) -> int | None:
    for name in CARBON_HEADERS:
        if name in headers:
            return headers.index(name)
    for i, h in enumerate(headers):
        if h.strip() == "탄소배출량":
            return i
    for i, h in enumerate(headers):
        if "탄소배출" in h or "배출량" in h:
            return i
    return None


def summarize(path: Path) -> dict:
    headers, data = load_rows(path)
    col = {h: i for i, h in enumerate(headers) if h}
    cidx = carbon_idx(headers)

    yms: set[tuple[int, int]] = set()
    regions: set[str] = set()
    carbon_vals: list[float] = []
    index_vals: list[float] = []
    industry_totals: dict[str, float] = defaultdict(float)

    for row in data:
        y = int(parse_number(row[col["year"]]))
        m = int(parse_number(row[col["month"]]))
        yms.add((y, m))
        sido = str(row[col["sido_nm"]]).strip()
        sgg = str(row[col["sgg_nm"]]).strip()
        regions.add(f"{sido} {sgg}")
        if cidx is not None:
            carbon_vals.append(parse_number(row[cidx]))
        index_vals.append(parse_number(row[col["탄소발자국 지수"]]))
        for name in INDUSTRY_COLUMNS:
            if name in col:
                industry_totals[name] += parse_number(row[col[name]])

    return {
        "file": path.name,
        "header_count": len(headers),
        "row_count": len(data),
        "headers": headers,
        "carbon_column": headers[cidx] if cidx is not None else None,
        "ym_min": min(yms),
        "ym_max": max(yms),
        "ym_count": len(yms),
        "region_count": len(regions),
        "carbon_sum": sum(carbon_vals),
        "carbon_min": min(carbon_vals) if carbon_vals else None,
        "carbon_max": max(carbon_vals) if carbon_vals else None,
        "index_min": min(index_vals) if index_vals else None,
        "index_max": max(index_vals) if index_vals else None,
        "industry_totals": dict(industry_totals),
        "missing_required": [c for c in REQUIRED if c not in col],
        "missing_industry": [c for c in INDUSTRY_COLUMNS if c not in col],
    }


def keyed_rows(path: Path) -> dict[tuple, dict]:
    headers, data = load_rows(path)
    col = {h: i for i, h in enumerate(headers) if h}
    cidx = carbon_idx(headers)
    out: dict[tuple, dict] = {}
    for row in data:
        key = (
            str(row[col["sido_nm"]]).strip(),
            str(row[col["sgg_nm"]]).strip(),
            int(parse_number(row[col["year"]])),
            int(parse_number(row[col["month"]])),
        )
        industries = {
            name: parse_number(row[col[name]])
            for name in INDUSTRY_COLUMNS
            if name in col
        }
        out[key] = {
            "carbon": parse_number(row[cidx]) if cidx is not None else None,
            "index": parse_number(row[col["탄소발자국 지수"]]),
            "industries": industries,
        }
    return out


def main() -> None:
    if not OLD.exists() or not NEW.exists():
        print("Missing file(s)", OLD.exists(), NEW.exists())
        sys.exit(1)

    old_s = summarize(OLD)
    new_s = summarize(NEW)

    print("=== SUMMARY ===")
    for s in (old_s, new_s):
        print(json.dumps({k: v for k, v in s.items() if k not in ("headers", "industry_totals")}, ensure_ascii=False, indent=2))
        print()

    old_headers = set(old_s["headers"])
    new_headers = set(new_s["headers"])
    only_old = sorted(old_headers - new_headers)
    only_new = sorted(new_headers - old_headers)

    print("=== HEADERS only in OLD (%d) ===" % len(only_old))
    print("\n".join(only_old[:50]))
    if len(only_old) > 50:
        print(f"... +{len(only_old)-50} more")

    print("\n=== HEADERS only in NEW (%d) ===" % len(only_new))
    print("\n".join(only_new[:80]))
    if len(only_new) > 80:
        print(f"... +{len(only_new)-80} more")

    print("\n=== CONVERTER COMPAT (new file) ===")
    print("missing_required:", new_s["missing_required"])
    print("missing_industry:", new_s["missing_industry"])
    print("carbon_column:", new_s["carbon_column"])

    old_rows = keyed_rows(OLD)
    new_rows = keyed_rows(NEW)
    only_old_keys = set(old_rows) - set(new_rows)
    only_new_keys = set(new_rows) - set(old_rows)
    common = set(old_rows) & set(new_rows)

    carbon_diff = 0
    index_diff = 0
    industry_diff = 0
    max_carbon_rel = 0.0
    sample_diffs: list[str] = []

    for key in common:
        o, n = old_rows[key], new_rows[key]
        oc, nc = o["carbon"] or 0, n["carbon"] or 0
        if abs(oc - nc) > 0.01:
            carbon_diff += 1
            if oc:
                max_carbon_rel = max(max_carbon_rel, abs(nc - oc) / abs(oc))
            if len(sample_diffs) < 5:
                sample_diffs.append(f"{key}: carbon {oc} -> {nc}")
        if abs(o["index"] - n["index"]) > 0.0001:
            index_diff += 1
        for name in INDUSTRY_COLUMNS:
            ov = o["industries"].get(name, 0)
            nv = n["industries"].get(name, 0)
            if abs(ov - nv) > 0.01:
                industry_diff += 1
                break

    print("\n=== ROW KEY DIFF ===")
    print("common:", len(common), "only_old:", len(only_old_keys), "only_new:", len(only_new_keys))
    print("carbon value diffs:", carbon_diff, "index diffs:", index_diff, "any industry diff rows:", industry_diff)
    print("max relative carbon diff:", round(max_carbon_rel, 6))
    for line in sample_diffs:
        print(" ", line)

    print("\n=== INDUSTRY TOTALS (period sum) ===")
    for name in INDUSTRY_COLUMNS:
        ot = old_s["industry_totals"].get(name, 0)
        nt = new_s["industry_totals"].get(name, 0)
        if abs(ot - nt) > 1:
            rel = abs(nt - ot) / ot if ot else float("inf")
            print(f"  {name}: old={ot:,.0f} new={nt:,.0f} delta={nt-ot:,.0f} rel={rel:.4%}")

    # Try running converter on new file
    print("\n=== CONVERTER TEST (new file) ===")
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        import convert_region_excel as conv

        wb = openpyxl.load_workbook(NEW, read_only=True, data_only=True)
        rows = conv.convert_workbook(wb.active)
        wb.close()
        print("SUCCESS rows:", len(rows))
        print("sample:", rows[0])
    except Exception as e:
        print("FAIL:", type(e).__name__, e)


if __name__ == "__main__":
    main()
